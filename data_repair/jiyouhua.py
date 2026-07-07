# -*- coding: utf-8 -*-
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import PatternFill

# ===========================================================
# =============== 工具函数 ===================================
# ===========================================================

def norm(s):
    """基础清洗"""
    return str(s).replace("\u200b", "").replace("\ufeff", "").strip() if s else ""

def strip_emoji(s: str) -> str:
    """去除 emoji、变体选择符、特殊符号"""
    if not s:
        return ""
    s = str(s)
    emoji_pattern = re.compile(
        "["
        "\U0001F300-\U0001F5FF"
        "\U0001F600-\U0001F64F"
        "\U0001F680-\U0001F6FF"
        "\U0001F700-\U0001F77F"
        "\U0001F780-\U0001F7FF"
        "\U0001F800-\U0001F8FF"
        "\U0001F900-\U0001F9FF"
        "\U0001FA00-\U0001FA6F"
        "\U0001FA70-\U0001FAFF"
        "\u200d"
        "\u2640-\u2642"
        "\u2600-\u2B55"
        "\ufe0f"
        "]+", flags=re.UNICODE)
    s = emoji_pattern.sub("", s)
    s = re.sub(r"[\u3000\u200B\u200C\u200D]", "", s)
    return s.strip()

def has_ji(s):
    """检测是否含“极”"""
    return "极" in str(s) if s else False

def normalize_time_str(t_raw: str) -> str:
    """统一时间格式为 yyyy-MM-dd HH:MM:SS"""
    t_raw = norm(t_raw)
    m = re.match(r"(\d{4}-\d{2}-\d{2})\s+(\d{1,2}):(\d{2}):(\d{2})$", t_raw)
    if not m:
        raise ValueError(f"无法解析时间: {t_raw}")
    date, hh, mm, ss = m.groups()
    hh = hh.zfill(2)
    return f"{date} {hh}:{mm}:{ss}"

def parse_txt(path: str) -> pd.DataFrame:
    """解析 TXT 为 DataFrame 并清洗 emoji、标准化时间"""
    pat_head = re.compile(r"(\d{4}-\d{2}-\d{2}\s+\d{1,2}:\d{2}:\d{2})\s+(.+)")
    recs = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.read().splitlines()
    n = len(lines)
    i = 0
    while i < n:
        line = lines[i].strip()
        m = pat_head.match(line)
        if m:
            t_raw, pid = m.groups()
            try:
                t_fmt = normalize_time_str(t_raw)
            except Exception:
                i += 1
                continue
            msg = lines[i + 1].strip() if (i + 1) < n else ""
            recs.append({
                "发言时间": t_fmt,
                "玩家ID": norm(pid),
                "玩家消息": strip_emoji(norm(msg))
            })
        i += 1
    df = pd.DataFrame(recs, columns=["发言时间","玩家ID","玩家消息"])
    print(f"📄 TXT 解析完成：{len(df)} 条，时间统一为 yyyy-MM-dd HH:MM:SS")
    return df

def find_top_left_cell(ws, row, col):
    """合并单元格安全写入"""
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(rng))
        if min_r <= row <= max_r and min_c <= col <= max_c:
            return ws.cell(row=min_r, column=min_c)
    return ws.cell(row=row, column=col)

# ===========================================================
# =============== 主逻辑 ====================================
# ===========================================================

def fix_extreme_by_rule(excel_path, txt_path, output_path, log_path):
    wb = load_workbook(excel_path)
    df_txt = parse_txt(txt_path)
    if not isinstance(df_txt, pd.DataFrame) or df_txt.empty:
        raise RuntimeError("❌ TXT 解析失败或为空，请检查 txt 文件格式。")

    fill_yellow = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    logs = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        data = ws.values
        try:
            cols = list(next(data))  # 第一行表头
        except StopIteration:
            continue

        df = pd.DataFrame(data, columns=cols)
        for c in ["话题簇","发言时间","玩家ID","玩家消息"]:
            if c not in df.columns:
                df[c] = ""

        print(f"🧩 开始处理 sheet：《{sheet}》 ...")

        # -------- ① 清理话题簇中的“极”、“极海听雷” --------
        for idx, val in enumerate(df["话题簇"]):
            if isinstance(val, str) and ("极" in val or "极海听雷" in val):
                new_val = val.replace("极海听雷", "").replace("极", "").strip()
                if new_val != val:
                    c_idx = cols.index("话题簇") + 1
                    cell = find_top_left_cell(ws, idx + 2, c_idx)
                    cell.value = new_val
                    cell.fill = fill_yellow
                    logs.append({
                        "sheet": sheet,
                        "行号": idx + 2,
                        "字段": "话题簇",
                        "原值": val,
                        "修正值": new_val,
                        "依据时间": ""
                    })

        # -------- ② 含极修正逻辑（增强版：发言时间+ID+消息修正） --------
        fix_cols = ["发言时间", "玩家ID", "玩家消息"]  # 话题簇不在第二轮覆盖

        for i, row in df.iterrows():
            t_raw = norm(row.get("发言时间"))
            pid   = strip_emoji(norm(row.get("玩家ID")))
            msg   = strip_emoji(norm(row.get("玩家消息")))
            topic = norm(row.get("话题簇"))

            # 四字段是否含极
            t_bad     = has_ji(t_raw)
            id_bad    = has_ji(pid)
            msg_bad   = has_ji(msg)
            topic_bad = has_ji(topic)

            if not (t_bad or id_bad or msg_bad or topic_bad):
                continue

            # 发言时间含极 → 先清洗成可匹配 TXT 的时间格式
            t_clean = t_raw.replace("极海听雷","").replace("极","极速蜗牛").strip()
            if not t_clean:
                continue

            # 匹配 TXT 原始记录
            cand = df_txt[df_txt["发言时间"] == t_clean]
            if cand.empty:
                continue
            ref = cand.iloc[0]

            # 修正“发言时间 / 玩家ID / 玩家消息”
            for col in fix_cols:
                old_val = norm(row.get(col))
                if has_ji(old_val):  # 只修含极的
                    new_val = norm(ref.get(col))
                    if new_val and new_val != old_val:
                        c_idx = cols.index(col) + 1
                        cell = find_top_left_cell(ws, i + 2, c_idx)
                        cell.value = new_val
                        cell.fill = fill_yellow
                        logs.append({
                            "sheet": sheet,
                            "行号": i + 2,
                            "字段": col,
                            "原值": old_val,
                            "修正值": new_val,
                            "依据时间": t_clean
                        })

        print(f"✅ Sheet 《{sheet}》 处理完成。")

    wb.save(output_path)
    print(f"💾 修正完成：{output_path}")

    if logs:
        pd.DataFrame(logs).to_excel(log_path, index=False)
        print(f"📝 修正日志生成：{log_path}")
    else:
        print("⚪ 未发现需要修正的字段。")

# ===========================================================
# =============== 入口 =======================================
# ===========================================================
if __name__ == "__main__":
    excel_path = r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\01231群_空值补齐V2.xlsx"
    txt_path   = r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\0123《欢迎来到地球》测试1群.txt"
    output_path= excel_path.replace(".xlsx", "修正.xlsx")
    log_path   = excel_path.replace(".xlsx", "_含极三态修正.xlsx")
    fix_extreme_by_rule(excel_path, txt_path, output_path, log_path)
    print("🎯 全流程执行完毕！")
