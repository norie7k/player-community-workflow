# -*- coding: utf-8 -*-
"""
对 02082群.xlsx 每个 sheet 的话题簇列（A列）做单元格合并：
相同话题簇名称的连续行合并为一个单元格，居中显示。
输出为新文件，不修改原文件。
"""

import openpyxl
from openpyxl.styles import Alignment
from pathlib import Path

# ==================== 配置 ====================

INPUT_FILE = Path(r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\02082群.xlsx")
OUTPUT_FILE = INPUT_FILE.parent / "02082群_已合并.xlsx"

TOPIC_COL = 1  # 话题簇所在列（A列）

# ==================== 主逻辑 ====================

def merge_topic_cells(ws):
    """
    对单个 sheet 的话题簇列做合并：
    找到连续相同值的行区间，执行 merge_cells
    """
    max_row = ws.max_row
    if max_row < 2:
        return 0

    merge_count = 0
    start_row = 2  # 跳过表头
    current_topic = ws.cell(row=2, column=TOPIC_COL).value

    for r in range(3, max_row + 2):  # +2 是为了处理最后一组
        if r <= max_row:
            cell_val = ws.cell(row=r, column=TOPIC_COL).value
        else:
            cell_val = None  # 哨兵值，强制结束最后一组

        if cell_val == current_topic and r <= max_row:
            continue
        else:
            # 当前组结束：start_row ~ r-1
            if r - 1 > start_row:
                ws.merge_cells(
                    start_row=start_row,
                    end_row=r - 1,
                    start_column=TOPIC_COL,
                    end_column=TOPIC_COL,
                )
                # 合并后设置居中对齐
                ws.cell(row=start_row, column=TOPIC_COL).alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                merge_count += 1

            # 开始新的一组
            start_row = r
            current_topic = cell_val

    return merge_count


def main():
    print(f"📂 读取文件: {INPUT_FILE.name}")
    wb = openpyxl.load_workbook(INPUT_FILE)

    print(f"   Sheets: {wb.sheetnames}\n")

    for sn in wb.sheetnames:
        ws = wb[sn]
        row_count = ws.max_row - 1  # 去掉表头
        merge_count = merge_topic_cells(ws)
        print(f"  ✅ [{sn}] {row_count} 行数据，合并了 {merge_count} 组话题簇")

    wb.save(OUTPUT_FILE)
    print(f"\n🎯 完成！已保存到: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
