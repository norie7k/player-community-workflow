# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from copy import copy

# ======================== 配置 =========================
SHEET_NAMES = ["体验反馈", "疑惑询问", "建议灵感", "情绪输出", "问题反馈"]
HEADERS     = ["话题簇", "发言时间", "玩家ID", "玩家消息"]

# 样式配置
FONT_NAME         = "微雅软黑"      # 若本机没有，可改 "微软雅黑"
HEADER_FONT_SIZE  = 16
BODY_FONT_SIZE    = 11
HEADER_FILL       = "FFDDEBF7"     # 浅蓝底
HEADER_ROW_HEIGHT = 24
COL_WIDTHS        = [16, 21, 30, 95]   # 对应各列宽度

# ======================== 工具函数 =========================
def copy_cell_style(src_cell, dest_cell):
    """复制单元格样式"""
    if src_cell and src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = copy(src_cell.number_format)
        dest_cell.protection = copy(src_cell.protection)
        dest_cell.alignment = copy(src_cell.alignment)

def set_sheet_format(ws):
    """应用统一样式：冻结首行、列宽、表头样式"""
    ws.freeze_panes = "A2"
    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    # 表头样式
    for c, text in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c)
        cell.value = text
        cell.font = Font(name=FONT_NAME, bold=True, size=HEADER_FONT_SIZE)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

def extract_topic_blocks(ws):
    """
    从一个 sheet 中提取所有「话题簇块」信息：
    每个块包含：
        {
            'topic': 话题簇名称,
            'rows': [[每列的值], ...],
            'merged': [(min_col, min_row, max_col, max_row)]
        }
    """
    topic_blocks = []
    topic_col = 1  # 话题簇列为第一列
    merges = list(ws.merged_cells.ranges)
    nrows, ncols = ws.max_row, ws.max_column

    merged_row_set = set()
    # 收集所有合并行的范围
    for rng in merges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        for r in range(min_row, max_row + 1):
            merged_row_set.add(r)
        # 仅处理话题簇列的合并区域
        if min_col == max_col == topic_col:
            topic = ws.cell(row=min_row, column=topic_col).value
            block_rows = []
            for r in range(min_row, max_row + 1):
                block_rows.append([ws.cell(row=r, column=c).value for c in range(1, ncols + 1)])
            topic_blocks.append({
                "topic": topic or "",
                "rows": block_rows,
                "merged": [(min_col, min_row, max_col, max_row)]
            })

    # 找出未合并的行（单独话题簇）
    for r in range(2, nrows + 1):
        if r not in merged_row_set:
            topic = ws.cell(row=r, column=topic_col).value or ""
            row_data = [[ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]]
            topic_blocks.append({
                "topic": topic,
                "rows": row_data,
                "merged": []  # 无合并结构
            })

    return topic_blocks


def copy_block_to(ws_dest, ws_src, block, start_row):
    """将整个话题簇块复制到目标 sheet（保留样式+合并结构）"""
    base_row = start_row
    ncols = ws_src.max_column

    # 🔹 如果该块无合并结构（单独行）
    if not block["merged"]:
        for i, row_values in enumerate(block["rows"], start=0):
            for j, val in enumerate(row_values, start=1):
                src = ws_src.cell(row=2, column=j)  # 从第2行取样式模板
                dest = ws_dest.cell(row=base_row + i, column=j, value=val)
                copy_cell_style(src, dest)
        return

    # 🔹 有合并结构的块
    min_col, min_row, max_col, max_row = block["merged"][0]
    for i, row_values in enumerate(block["rows"], start=0):
        for j, val in enumerate(row_values, start=1):
            src = ws_src.cell(row=min_row + i, column=j)
            dest = ws_dest.cell(row=base_row + i, column=j, value=val)
            copy_cell_style(src, dest)

    # 重建合并结构
    for (min_col, _, max_col, max_row) in block["merged"]:
        height = max_row - min_row
        ws_dest.merge_cells(
            start_row=base_row,
            end_row=base_row + height,
            start_column=min_col,
            end_column=max_col
        )

# ======================== 主逻辑 =========================
def merge_excels_keep_blocks_sorted(excel_files, output_path):
    print(f"🧩 合并 {len(excel_files)} 个 Excel，按话题簇块排序，保留原合并结构...")

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in SHEET_NAMES:
        print(f"\n🔧 处理 Sheet：《{sheet_name}》 ...")
        all_blocks = []
        template_ws = None

        # 收集所有文件中该 Sheet 的话题簇块
        for path in excel_files:
            wb = openpyxl.load_workbook(path)
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ {path} 缺少 Sheet：《{sheet_name}》 ，跳过。")
                continue
            ws = wb[sheet_name]
            if template_ws is None:
                template_ws = ws
            blocks = extract_topic_blocks(ws)
            all_blocks.extend(blocks)

        # 按话题簇名称升序排序（空值排最后）
        all_blocks.sort(key=lambda b: (b["topic"] == "", b["topic"] or ""))

        # 写入新工作簿
        ws_new = wb_out.create_sheet(title=sheet_name)
        set_sheet_format(ws_new)

        current_row = 2
        for block in all_blocks:
            copy_block_to(ws_new, template_ws, block, current_row)
            current_row += len(block["rows"])

        print(f"✅ Sheet 《{sheet_name}》 完成：共 {len(all_blocks)} 个话题簇块。")

    wb_out.save(output_path)
    print(f"\n🎯 所有 Sheet 合并完成，保留合并格式输出：{output_path}")

# ======================== 执行入口 =========================


# ======================== 执行入口 =========================
if __name__ == "__main__":
    excel_files = [
        r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\02232群.xlsx",
        r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\【地球】_「Beta19_立春测试」_玩家发言分类「2群」_260204~0213（持续更新）.xlsx"
        
    ]
    output_path = r"E:\项目\玩家社群分析智能体\玩家发言分类（供研发侧）\【地球】_「Beta19_立春测试」_玩家发言分类「2群」_260204~0223（持续更新）.xlsx"
    merge_excels_keep_blocks_sorted(excel_files, output_path)



 


  