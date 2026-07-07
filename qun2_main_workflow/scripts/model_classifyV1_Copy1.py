
from __future__ import annotations
import json, time, typing as T
import pandas as pd
import requests
from pathlib import Path
from typing import List, Dict, Any
import re, json, unicodedata
import json
# --- openpyxl 样式/工具 ---
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle

import re

################模型调用，出结果###################

def load_system_prompt(path: Path) -> str:
    if not path.exists():
        raise FileNotFoundError(f"Prompt file not found: {path}")
    return path.read_text(encoding="utf-8")

def build_user_prompt_filter(json_lines: T.List[str]) -> str:
    # 模型#1：筛掉非游戏相关，只输出相关 JSON 行（原样）
    return (
        "以下是若干玩家/客服/研发的发言记录，请根据系统提示中规则，"
        "判断哪些是【与游戏内容相关】的发言，保留这些 JSON 行，不相关的忽略。"
        "请仅输出【相关发言的原始 JSON 行】，严格保持格式不变。\n\n"
        "【输入】：\n" + "\n".join(json_lines)
    )

def build_user_prompt_classify2(jsonl_block: str) -> str:
    # 模型#3：对已筛选的相关 JSON 行（原样）进行二级标签分类
    return (
       "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请仅输出【 JSON 行】，\n\n"
        "【输入】：\n" + jsonl_block
    )
    
def build_user_prompt_classify(jsonl_block: str) -> str:
    # 模型#2：对已筛选的相关 JSON 行（原样）进行分类，追加“意图分类”键
    return (
       "以下是输入数据（JSONL 格式，每行一个发言对象）：\n\n"
        "请仅输出【 JSON 行】，\n\n"
        "【输入】：\n" + jsonl_block
    )

def build_user_prompt_cluster_correct(clustered_jsonl: str, whitelist: List[Dict]) -> str:
    whitelist_text = "[当前白名单为空，暂无参考命名]" if not whitelist else "\n".join(
        json.dumps(x, ensure_ascii=False) for x in whitelist
    )
    return (
        "你是一位“话题簇命名校正专家”。对照白名单统一命名；不匹配则保留原名。\n"
        "仅输出字段：发言日期、发言时间、玩家ID、玩家消息、分类标签、话题簇、话题簇描述。\n"
        "命中白名单时：话题簇 = 白名单名称；话题簇描述 = 白名单相关描述。\n"
        "未命中白名单时：话题簇保持输入值；话题簇描述沿用输入（若无则输出空字符串）。\n\n"
        "【发言】：\n" + clustered_jsonl + "\n\n"
        "【白名单】：\n" + whitelist_text + "\n"
    )


def call_ark_chat_completions(
    api_url: str,
    api_key: str,
    model: str,
    system_prompt: str,
    user_prompt: str,
    temperature: float = 0.3,
    max_tokens: int = 32700,
    timeout: int = 600,
    retries: int = 2,
) -> str:
    headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(api_url, headers=headers, json=payload, timeout=timeout)
            if resp.status_code != 200:
                raise RuntimeError(f"HTTP {resp.status_code}: {resp.text}")
            data = resp.json()
            return data["choices"][0]["message"]["content"]
        except Exception as e:
            last_err = e
            time.sleep(1.2 * (attempt + 1))
    raise RuntimeError(f"Ark API 调用失败: {last_err}")

def extract_valid_json_lines(text: str) -> T.List[str]:
    """
    把模型输出里的纯 JSON 行提取出来（鲁棒处理）：
    - 逐行判断：以 { 开头 且 以 } 结尾，则认为是一个 JSON 对象行
    - 也能容忍前后多余空行或解释文字（会被忽略）
    """
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("{") and s.endswith("}"):
            lines.append(s)
    return lines






def jsonl_to_dataframe_with_intent(jsonl_text: str) -> pd.DataFrame:
    """
    将模型#3输出（JSONL，每行一个JSON）转 df。
    目标列：["话题簇","发言时间","玩家ID","玩家消息","一级分类"]

    新增能力：
    - 支持输入同时包含「发言日期」「发言时间」两个字段；
    - 自动合并为统一的「发言时间」（格式：%Y-%m-%d %H:%M:%S）。
    """
    # 取出纯 JSON 行
    try:
        pure_lines = extract_valid_json_lines(jsonl_text)
    except NameError:
        pure_lines = [ln.strip() for ln in (jsonl_text or "").splitlines() if ln.strip()]

    if not pure_lines:
        return pd.DataFrame(columns=["话题簇","发言时间","玩家ID","玩家消息","一级分类"])

    rows = []
    for line in pure_lines:
        try:
            rows.append(json.loads(line))
        except Exception:
            continue
    if not rows:
        return pd.DataFrame(columns=["话题簇","发言时间","玩家ID","玩家消息","一级分类"])

    df = pd.DataFrame(rows)

    # 键名兼容
    df = df.rename(columns={
        "玩家 ID": "玩家ID",
        "意图分类": "一级分类",
        "分类标签": "一级分类",
    })

    # 补列
    for c in ["发言日期","发言时间","玩家ID","玩家消息","一级分类","话题簇"]:
        if c not in df.columns:
            df[c] = pd.NA

    # 一级分类数值化（可选）
    df["一级分类"] = pd.to_numeric(df["一级分类"], errors="coerce")

    # 话题簇统一为字符串（如果是列表则用“、”拼接）
    def _topic_to_str(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return ""
        if isinstance(x, list):
            return "、".join(str(i) for i in x if i is not None)
        return str(x)
    df["话题簇"] = df["话题簇"].apply(_topic_to_str).str.strip()

    # === 合并“发言日期 + 发言时间” => 标准“发言时间” ===
    def _combine_dt(row):
        d = str(row.get("发言日期") or "").strip()
        t = str(row.get("发言时间") or "").strip()

        # 只有日期或只有时间的情况也容错
        if d and t:
            s = f"{d} {t}"
        elif d:
            s = d
        else:
            s = t

        # 统一分隔符，修正仅到分钟的时间补秒
        s = re.sub(r"[/.]", "-", s)
        if re.fullmatch(r"\d{1,2}:\d{2}", t):  # e.g. 14:03
            s = f"{d} {t}:00"

        ts = pd.to_datetime(s, errors="coerce")
        return ts

    ts = df.apply(_combine_dt, axis=1)

    # 输出为标准格式字符串；解析失败保持 NaN
    df["发言时间"] = ts.dt.strftime("%Y-%m-%d %H:%M:%S")
    df.loc[ts.isna(), "发言时间"] = pd.NA

    # 只保留目标列顺序
    return df[["话题簇","发言时间","玩家ID","玩家消息","一级分类"]]

##########################话提簇数据库################################
def load_whitelist(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        return [json.loads(line) for line in f if line.strip()]

def extract_clusters_from_output(output_text: str) -> list[dict]:
    import json, re
    if not output_text:
        return []
    s = output_text.strip()
    s = re.sub(r"^```[a-zA-Z0-9]*\s*", "", s)
    s = re.sub(r"\s*```$", "", s)
    results, seen = [], set()
    def _push(obj):
        cluster = (obj.get("话题簇") or "").strip()
        desc = (obj.get("话题簇描述") or obj.get("描述") or "").strip()
        if cluster and cluster not in seen:
            results.append({"话题簇名称": cluster, "相关描述": desc})
            seen.add(cluster)
    try:
        whole = json.loads(s)
        if isinstance(whole, dict):
            _push(whole); return results
        if isinstance(whole, list):
            for it in whole:
                if isinstance(it, dict): _push(it)
            return results
    except Exception:
        pass
    buf, depth = [], 0
    for ch in s:
        if ch == '{': depth += 1
        if depth > 0: buf.append(ch)
        if ch == '}':
            depth -= 1
            if depth == 0 and buf:
                block = ''.join(buf).strip(); buf = []
                try:
                    obj = json.loads(block)
                    if isinstance(obj, dict): _push(obj)
                except Exception:
                    pass
    return results


def update_and_save_whitelist(path: Path, current: list[dict], new_items: list[dict]) -> list[dict]:
    existing_names = {item["话题簇名称"] for item in current}
    added = [item for item in new_items if item["话题簇名称"] not in existing_names]

    if added:
        with open(path, "a", encoding="utf-8") as f:
            for item in added:
                f.write(json.dumps(item, ensure_ascii=False) + "\n")
        current.extend(added)
        print(f"✅ 新增 {len(added)} 条话题簇至白名单")
    else:
        print("⚪ 无新增话题簇")

    return current


##########################导入Excel格式要求###########################

from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
import unicodedata
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

########################## Excel 结构与样式 ###########################
SHEET_NAMES = ["体验反馈", "疑惑询问", "建议灵感", "情绪输出", "问题反馈"]
# 新增“二级标签”列
HEADERS = [ "话题簇","发言时间", "玩家ID", "玩家消息"]

# 样式配置
FONT_NAME = "微雅软黑"         # 若本机没有，可改 "微软雅黑"
HEADER_FONT_SIZE = 16
BODY_FONT_SIZE   = 11
HEADER_FILL      = "FFDDEBF7"     # 蓝底
HEADER_ROW_HEIGHT = 24
COL_WIDTHS = [ 16,21, 30, 95]   # 时间 / 玩家ID / 玩家消息 / 二级标签

CA_TO_SHEET = {1:"体验反馈", 2:"疑惑询问", 3:"建议灵感", 4:"情绪输出", 5:"问题反馈"}


def _ensure_named_style(wb) -> str:
    """确保正文 NamedStyle 存在；返回 style 名称。"""
    style_name = "BodyStyle"
    if style_name in wb.named_styles:
        return style_name
    thin = Side(style="thin", color="000000")
    body_style = NamedStyle(name=style_name)
    body_style.font = Font(name=FONT_NAME, size=BODY_FONT_SIZE)
    body_style.alignment = Alignment(vertical="center", wrap_text=True)
    body_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.add_named_style(body_style)
    return style_name


def create_intent_excel_styled(filename: str):
    """初始化工作簿：5 个 sheet，表头+列宽+冻结+表头样式+正文样式注册"""
    wb = Workbook()
    ws0 = wb.active
    ws0.title = SHEET_NAMES[0]

    for idx, name in enumerate(SHEET_NAMES):
        ws = wb[name] if idx == 0 else wb.create_sheet(title=name)

        # 表头
        ws.append(HEADERS)

        # 列宽
        for col_idx, w in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # 冻结首行
        ws.freeze_panes = "A2"

        # 表头样式
        for c in range(1, len(HEADERS)+1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name=FONT_NAME, size=HEADER_FONT_SIZE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)

        ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    _ensure_named_style(wb)
    wb.save(filename)
    print(f"✅ 创建并套样式：{filename}")


def _open_or_create_excel(excel_path: str):
    """没有文件就创建并预置格式；返回 (wb, created_flag)"""
    p = Path(excel_path)
    if not p.exists():
        create_intent_excel_styled(excel_path)
        wb = load_workbook(excel_path)
        return wb, True
    wb = load_workbook(excel_path)
    _ensure_named_style(wb)
    return wb, False


########################## 数据规范化 ###########################
def _normalize_records(records: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    输入：形如
    [{"一级分类":2,"二级标签":["殖装保留","保留猜测"],"发言时间":"2025-08-06 17:25:36",
      "玩家 ID":"，(1272414483)","玩家消息":"能保留殖装应该"}, ...]
    输出：DataFrame，列 = ["一级分类","发言时间","玩家ID","玩家消息","二级标签"]，
    其中“二级标签”为单个字符串（若原本是列表则 explode）
    """
    if not records:
        return pd.DataFrame(columns=["一级分类", "发言时间", "玩家ID", "玩家消息", "话题簇"])

    df = pd.DataFrame(records)

    # 兼容 键名差异：玩家 ID vs 玩家ID
    if "玩家ID" not in df.columns and "玩家 ID" in df.columns:
        df = df.rename(columns={"玩家 ID": "玩家ID"})

    # 缺列补空
    for col in ["一级分类", "发言时间", "玩家ID", "玩家消息", "话题簇"]:
        if col not in df.columns:
            df[col] = pd.NA

    # 统一“二级标签”为列表
    def _to_list(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return []
        if isinstance(x, list):
            return x
        # 字符串就当成单标签
        return [str(x)]

    df["话题簇"] = df["话题簇"].apply(_to_list)

    # explode 成单标签一行；若原本为空列表，将生成空行，先过滤
    df = df.explode("话题簇", ignore_index=True)
    df["话题簇"] = df["话题簇"].fillna("").astype(str)

    # 一级分类数值化 & 过滤有效 sheet
    df["一级分类"] = pd.to_numeric(df["一级分类"], errors="coerce")
    df = df[df["一级分类"].isin(CA_TO_SHEET.keys())]
    # df = df.rename(columns={"二级标签": "话题簇"})

    # 只保留需要列、并按既定顺序
    return df[["话题簇","一级分类", "发言时间", "玩家ID", "玩家消息" ]]


########################## 写入 + 按二级标签热度排序 ###########################
def _rewrite_sheet_sorted_by_tag(wb, sheet_name: str, new_rows_df: pd.DataFrame):
    """
    合并历史+新增后，按：
      1) 二级标签 升序（空标签放最后）
      2) 发言时间 升序
    重写正文区域（A2:D*）。
    """
    ws = wb[sheet_name]
    body_style_name = _ensure_named_style(wb)

    # 读取既有正文（A2:D*）
    existing = []
    if ws.max_row >= 2:
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row,
                              min_col=1, max_col=len(HEADERS), values_only=True):
            if r is None or all(x in (None, "") for x in r):
                continue
            existing.append(r)

    cols = HEADERS  # ["发言时间","玩家ID","玩家消息","二级标签"]
    df_exist = pd.DataFrame(existing, columns=cols) if existing else pd.DataFrame(columns=cols)

    # 合并历史+新增
    df_all = pd.concat([df_exist, new_rows_df[cols]], ignore_index=True)
    if df_all.empty:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        return

    # 规范化：标签去首尾空白；构造“是否空标签”标记；解析时间
    df_all["话题簇"] = df_all["话题簇"].astype(str).str.strip()
    df_all["_tag_blank"] = df_all["话题簇"].eq("") | df_all["话题簇"].isna()
    df_all["_ts"] =pd.to_datetime(df_all["发言时间"], format="%Y-%m-%d %H:%M:%S", errors="coerce")

    # 关键排序：非空标签在前 -> 标签升序 -> 时间升序（无法解析的时间排后，通过原字符串兜底）
    df_all = df_all.sort_values(
        by=["_tag_blank", "话题簇", "_ts", "发言时间"],
        ascending=[True, True, True, True],
        kind="mergesort"  # 稳定排序
    ).drop(columns=["_tag_blank", "_ts"])

    # 清正文并写回
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row in df_all.itertuples(index=False, name=None):
        ws.append(row)

    # 套正文样式
    if ws.max_row >= 2:
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row,
                              min_col=1, max_col=len(HEADERS)):
            for cell in r:
                cell.style = body_style_name



def append_json_to_excel_by_cat_and_tag(records: List[Dict[str, Any]], excel_path: str):
    """
    主入口：
    - 解析/规范化输入记录
    - 按“一级分类”分发到 sheet
    - 每个 sheet 写入后，对“所有内容（含历史+新增）”按“二级标签”热度降序重排
    """
    df = _normalize_records(records)
    if df.empty:
        return

    wb, _ = _open_or_create_excel(excel_path)

    # 按一级分类分组，每个组在各自 sheet 内完成“合并 + 重排 + 重写”
    for ca, grp in df.groupby("一级分类", sort=False):
        sheet = CA_TO_SHEET.get(int(ca))
        if not sheet or sheet not in wb.sheetnames:
            continue

        # 仅保留写入列顺序
        body_df = grp[["话题簇","发言时间", "玩家ID", "玩家消息" ]].copy()

        _rewrite_sheet_sorted_by_tag(wb, sheet, body_df)

    wb.save(excel_path)





# ----------------------------Excel处理---------------------



# ========== 列号常量（1-based，对应你的新表头顺序） ==========
TOPIC_COL = 1      # A 列：话题簇
TIME_COL  = 2      # B 列：发言时间
USER_COL  = 3      # C 列：玩家ID
MSG_COL   = 4      # D 列：玩家消息

# ========== 表格样式（如果你需要初始化/保证正文样式存在） ==========
FONT_NAME = "微雅软黑"      # 没有就换成“微软雅黑”
BODY_FONT_SIZE = 11

def _ensure_named_style(wb) -> str:
    """确保正文 NamedStyle 存在；返回 style 名称。"""
    style_name = "BodyStyle"
    if style_name in wb.named_styles:
        return style_name
    thin = Side(style="thin", color="000000")
    body_style = NamedStyle(name=style_name)
    body_style.font = Font(name=FONT_NAME, size=BODY_FONT_SIZE)
    body_style.alignment = Alignment(vertical="center", wrap_text=True)
    body_style.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.add_named_style(body_style)
    return style_name

# ========== 读前：拆并并回填“话题簇”列 ==========
def _pre_unmerge_and_fill_topic(ws):
    to_process = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col == TOPIC_COL and rng.max_col == TOPIC_COL:
            to_process.append(rng)
    for rng in to_process:
        top_val = ws.cell(row=rng.min_row, column=TOPIC_COL).value
        for r in range(rng.min_row, rng.max_row + 1):
            ws.cell(row=r, column=TOPIC_COL).value = top_val
        ws.unmerge_cells(range_string=str(rng))

# ========== 文本规范化（去零宽/全角空格/多空白） ==========
_ZW_RE = re.compile(r'[\u200b-\u200f\u202a-\u202e\u2060-\u206f\ufeff]')
def _norm_topic(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = _ZW_RE.sub("", s).replace("\u3000", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s

# ========== 彻底删除图片/图表/绘图关系 ==========
def _remove_all_drawings(ws):
    # 清空图片与图表集合
    if hasattr(ws, "_images"):
        ws._images = []
    if hasattr(ws, "_charts"):
        ws._charts = []
    # 删除 drawing 关系（避免残留）
    try:
        rels = getattr(ws, "_rels", None)
        if rels:
            to_del = [rid for rid, rel in rels.items()
                      if "drawing" in getattr(rel, "type", "")]
            for rid in to_del:
                del rels[rid]
    except Exception:
        pass
    # 清空 _drawing 句柄
    if hasattr(ws, "_drawing"):
        ws._drawing = None

# ========== 清旧合并（仅话题簇列） ==========
def _clear_topic_merges(ws):
    try:
        for rng in list(ws.merged_cells.ranges):
            if rng.min_col == TOPIC_COL and rng.max_col == TOPIC_COL:
                ws.unmerge_cells(str(rng))
    except Exception:
        pass

# ========== 纯图片行判定 & 去重表情 ==========
_IMG_PAT = re.compile(r'^\s*(?:\[(?:图片|表情|图像)[^\]]*\]\s*)+$')
def _is_pure_image_msg(text: str) -> bool:
    if text is None:
        return False
    return bool(_IMG_PAT.match(str(text).strip()))

def _strip_trailing_flag(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r'(?:\s*🖼️)+$', '', str(s).rstrip())

# ========== 分段：同簇且相邻时间间隔≤gap，返回 df 行号段 ==========
def _iter_topic_runs(df: pd.DataFrame, gap_minutes: int, nat_policy: str = "skip"):
    """
    df: 必须包含列 ["发言时间","话题簇"]，且已按『话题簇→时间』排过序。
    返回: [(start_idx, end_idx, topic_norm), ...] —— 注意是 df 的 0-based 行号。
    """
    def _to_ts(x):
        try:
            return pd.to_datetime(x, errors="coerce")
        except Exception:
            return pd.NaT

    n = len(df)
    if n == 0:
        return []

    segs = []
    start = 0
    topic = str(df.iloc[0]["话题簇"] or "")
    last_ts = _to_ts(df.iloc[0]["发言时间"])

    for i in range(1, n):
        cur_topic = str(df.iloc[i]["话题簇"] or "")
        cur_ts = _to_ts(df.iloc[i]["发言时间"])

        same_topic = (cur_topic == topic and cur_topic != "")
        cont = False

        if same_topic:
            if pd.notna(cur_ts) and pd.notna(last_ts):
                cont = (cur_ts - last_ts) <= pd.Timedelta(minutes=gap_minutes)
                if cont:
                    last_ts = cur_ts
            elif pd.isna(cur_ts):
                # NaT：skip=并入但不更新基准；break=直接断段
                cont = (nat_policy != "break")
            else:  # last_ts NaT, cur_ts 可解析
                if nat_policy == "break":
                    cont = False
                else:
                    cont = True
                    last_ts = cur_ts
        else:
            cont = False

        if not cont:
            segs.append((start, i - 1, topic))
            start = i
            topic = cur_topic
            last_ts = cur_ts if pd.notna(cur_ts) else pd.NaT

    segs.append((start, n - 1, topic))
    # 过滤空 topic
    return [(s, e, t) for (s, e, t) in segs if (t or "").strip()]

          
# ========== 分段：同簇且相邻时间间隔≤gap，返回 df 行号段 ==========
def _iter_topic_runs(df: pd.DataFrame, gap_minutes: int, nat_policy: str = "skip"):
    """
    df: 必须包含列 ["发言时间","话题簇"]，且已按『话题簇→时间』排过序。
    返回: [(start_idx, end_idx, topic_norm), ...] —— 注意是 df 的 0-based 行号。
    """
    def _to_ts(x):
        try:
            return pd.to_datetime(x, errors="coerce")
        except Exception:
            return pd.NaT

    n = len(df)
    if n == 0:
        return []

    segs = []
    start = 0
    topic = str(df.iloc[0]["话题簇"] or "")
    last_ts = _to_ts(df.iloc[0]["发言时间"])

    for i in range(1, n):
        cur_topic = str(df.iloc[i]["话题簇"] or "")
        cur_ts = _to_ts(df.iloc[i]["发言时间"])

        same_topic = (cur_topic == topic and cur_topic != "")
        cont = False

        if same_topic:
            if pd.notna(cur_ts) and pd.notna(last_ts):
                cont = (cur_ts - last_ts) <= pd.Timedelta(minutes=gap_minutes)
                if cont:
                    last_ts = cur_ts
            elif pd.isna(cur_ts):
                # NaT：skip=并入但不更新基准；break=直接断段
                cont = (nat_policy != "break")
            else:  # last_ts NaT, cur_ts 可解析
                if nat_policy == "break":
                    cont = False
                else:
                    cont = True
                    last_ts = cur_ts
        else:
            cont = False

        if not cont:
            segs.append((start, i - 1, topic))
            start = i
            topic = cur_topic
            last_ts = cur_ts if pd.notna(cur_ts) else pd.NaT

    segs.append((start, n - 1, topic))
    # 过滤空 topic
    return [(s, e, t) for (s, e, t) in segs if (t or "").strip()]

# ========== 核心：排序→重写→切段→段内排序→段间排序→合并→标🖼️→清图 ==========
def _sort_merge_flag(
    ws,
    gap_minutes: int = 15,
    nat_policy: str = "skip",          # "skip": NaT 并入不更新基准；"break": NaT 直接断段
    fill_nat_in_topic: bool = True,    # 同簇内对少量 NaT 做前后填补
    dump_bad_ts: bool = False
):
    _pre_unmerge_and_fill_topic(ws)

    # 读正文
    rows = []
    if ws.max_row >= 2:
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
            if r is None or all(x in (None, "") for x in r):
                continue
            rows.append(r)
    df = pd.DataFrame(rows, columns=["话题簇","发言时间","玩家ID","玩家消息"])
    if df.empty:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        _clear_topic_merges(ws); _remove_all_drawings(ws)
        return

    # 规范 & 强韧解析时间
    df["话题簇"] = df["话题簇"].apply(_norm_topic)
    col = df["发言时间"].astype(str).str.replace(r"[/.]", "-", regex=True).str.strip()
    ts  = pd.to_datetime(col, errors="coerce")
    bad = ts.isna()
    if bad.any():
        ts2 = pd.to_datetime(df.loc[bad, "发言时间"], format="%Y-%m-%d %H:%M:%S", errors="coerce")
        ts.loc[bad] = ts2
    df["_ts"] = ts

    if dump_bad_ts and df["_ts"].isna().any():
        df.loc[df["_ts"].isna(), ["话题簇","发言时间","玩家ID","玩家消息"]].to_excel(
            f"bad_ts_{ws.title}.xlsx", index=False
        )

    if fill_nat_in_topic:
        def _safe_fill(s: pd.Series) -> pd.Series:
            if s.notna().any():
                return s.ffill().bfill()
            return s
        df["_ts"] = df.groupby("话题簇")["_ts"].transform(_safe_fill)

    # <<< 新增：若原“发言时间”为空且 _ts 有值，用 _ts 回填为标准字符串，避免写表空白
    _time_is_empty = df["发言时间"].isna() | (df["发言时间"].astype(str).str.strip() == "")
    mask_fill = _time_is_empty & df["_ts"].notna()
    if mask_fill.any():
        df.loc[mask_fill, "发言时间"] = df.loc[mask_fill, "_ts"].dt.strftime("%Y-%m-%d %H:%M:%S")

    # <<< 新增：玩家消息 NaN → ""（避免 openpyxl 写 None → Excel 空）
    df["玩家消息"] = df["玩家消息"].astype(object).where(df["玩家消息"].notna(), "")

    # 基础排序：话题簇→时间，便于切段稳定（同簇内时间升序）
    df = df.sort_values(by=["话题簇", "_ts", "发言时间"],
                        ascending=[True, True, True],
                        kind="mergesort")

    # ① 切段
    runs = _iter_topic_runs(df[["发言时间","话题簇"]].copy(), gap_minutes, nat_policy)

    # ② 段内排序；③ 段与段之间按“话题簇 + 段首时间”排序（确保同簇整体时间升序）
    parts = []
    run_meta = []   # <<< 修改：[(idx_in_parts, topic, start_ts)]
    for (s, e, t) in runs:
        seg = df.iloc[s:e+1].copy()
        seg = seg.sort_values(by=["_ts","发言时间"], ascending=[True, True], kind="mergesort")
        parts.append(seg)
        seg_ts = pd.to_datetime(seg["发言时间"], errors="coerce")
        start_ts = seg_ts.iloc[0] if not seg_ts.empty else pd.NaT
        run_meta.append((len(parts)-1, t, start_ts))  # <<< 记录话题簇

    # <<< 修改：段间排序键 (topic, isNaT, start_ts)；NaT 段放最后
    run_meta.sort(key=lambda x: (x[1], pd.isna(x[2]), x[2]))
    df_out = pd.concat([parts[i] for (i, _topic, _ts0) in run_meta], ignore_index=True)

    # 重写 Excel（按重排后的顺序）
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    style = _ensure_named_style(ws.parent)
    for row in df_out.drop(columns=["_ts"]).itertuples(index=False, name=None):
        ws.append(row)
    for rr in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in rr:
            cell.style = style

    # 基于新顺序再切一次段 → 用于合并与标记
    runs_new = _iter_topic_runs(
        pd.DataFrame({
            "发言时间": [ws.cell(row=r, column=TIME_COL).value  for r in range(2, ws.max_row+1)],
            "话题簇":   [ws.cell(row=r, column=TOPIC_COL).value  for r in range(2, ws.max_row+1)],
        }),
        gap_minutes, nat_policy
    )

    # 合并 A 列（话题簇）
    _clear_topic_merges(ws)
    for (s, e, _t) in runs_new:
        r1, r2 = s + 2, e + 2
        if r2 > r1:
            ws.merge_cells(start_row=r1, start_column=TOPIC_COL, end_row=r2, end_column=TOPIC_COL)

    # 标记 🖼️：段内出现纯图片行 → D 列尾部 + 段首 A 列 + 🖼️
    for (s, e, _t) in runs_new:
        r1, r2 = s + 2, e + 2
        has_img_only = False
        for r in range(r1, r2 + 1):
            msg = ws.cell(row=r, column=MSG_COL).value
            if _is_pure_image_msg(msg):
                has_img_only = True
                ws.cell(row=r, column=MSG_COL).value = _strip_trailing_flag(str(msg)) + " 🖼️"  # <<< str()
        if has_img_only:
            tval = ws.cell(row=r1, column=TOPIC_COL).value
            ws.cell(row=r1, column=TOPIC_COL).value = _strip_trailing_flag(str(tval)) + " 🖼️"   # <<< str()

    # 移除所有 shape（小图片图标）
    _remove_all_drawings(ws)



# ========== 入口：处理整个工作簿 ==========
def postprocess_excel_by_topic(excel_path: str, gap_minutes: int = 15, nat_policy: str = "skip"):
    wb = load_workbook(excel_path, data_only=True)
    _ensure_named_style(wb)
    for name in wb.sheetnames:
        _sort_merge_flag(wb[name], gap_minutes=gap_minutes, nat_policy=nat_policy)
    wb.save(excel_path)
    print(f"✅ 已完成后处理：{excel_path}（gap={gap_minutes}min, NaT策略={nat_policy}）")

